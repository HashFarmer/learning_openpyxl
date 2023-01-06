from openpyxl import load_workbook

all_data = load_workbook("all_data.xlsx")

all_data_sheet_names = all_data.sheetnames

target_data = load_workbook("target_data.xlsx")


# 物业管理-应付职工薪酬[DL18],应交增值税[DO18],营业利润[CW18]，本年折旧[AC18]，税金及附加[BJ18]，投资收益[CN18]

for target_position,source_position in [["B6:B24","DL18"],["C6:C24","DO18"],["D6:D24","CW18"],["E6:E24","AC18"],["F6:F24","BJ18"],["G6:G24","CN18"]]:
	i = 0
	for cell_column in target_data["Sheet1"][target_position]:
		for cell in cell_column:
			cell.value = all_data[all_data_sheet_names[i]][source_position].value
			i+=1


# 房地产中介-应付职工薪酬[DL19],应交增值税[DO19],营业利润[CW19]，本年折旧[AC19]，税金及附加[BJ19]，投资收益[CN19]

for target_position,source_position in [["H6:H24","DL19"],["I6:I24","DO19"],["J6:J24","CW19"],["K6:K24","AC19"],["L6:L24","BJ19"],["M6:M24","CN19"]]:
	i = 0
	for cell_column in target_data["Sheet1"][target_position]:
		for cell in cell_column:
			cell.value = all_data[all_data_sheet_names[i]][source_position].value
			i+=1


# 房地产经营租赁-应付职工薪酬[DL20],应交增值税[DO20],营业利润[CW20]，本年折旧[AC20]，税金及附加[BJ20]，投资收益[CN20]

for target_position,source_position in [["N6:N24","DL20"],["O6:O24","DO20"],["P6:P24","CW20"],["Q6:Q24","AC20"],["R6:R24","BJ20"],["S6:S24","CN20"]]:
	i = 0
	for cell_column in target_data["Sheet1"][target_position]:
		for cell in cell_column:
			cell.value = all_data[all_data_sheet_names[i]][source_position].value
			i+=1


# 其他房地产业-应付职工薪酬[DL21],应交增值税[DO21],营业利润[CW21]，本年折旧[AC21]，税金及附加[BJ21]，投资收益[CN21]

for target_position,source_position in [["T6:T24","DL21"],["U6:U24","DO21"],["V6:V24","CW21"],["W6:W24","AC21"],["X6:X24","BJ21"],["Y6:Y24","CN21"]]:
	i = 0
	for cell_column in target_data["Sheet1"][target_position]:
		for cell in cell_column:
			cell.value = all_data[all_data_sheet_names[i]][source_position].value
			i+=1


target_data.save("target_data.xlsx")


