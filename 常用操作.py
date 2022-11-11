from openpyxl import Workbook, load_workbook

### 关于工作簿

# 创建、保存一个空白工作簿
wb = Workbook() 
wb.save("wb_openpyxl_3.xlsx") #所有针对工作簿的操作在此之前，而且这句是必须，否则前面的操作无法保存
wb.close()

#加载指定的工作簿
wb = load_workbook('hello_world.xlsx')


print(wb.encoding, end='\n\n')  # 获取文档的字符集编码
print(wb.properties)  # 获取文档的元数据如标题，创建者，创建日期等




### 关于工作表

#创建新的工作表
sheet_a = wb.create_sheet('123') #在所有工作表后面
#
sheet_b = wb.create_sheet('test',0) #在指定位置创建工作表

#获取所有的工作表名称
sheet_names = wb.get_sheet_names()
print(sheet_names)

#选择指定的工作表
# 方法1
ws = wb.get_sheet_by_name(sheet_names[0])
# 方法2
ws2 = wb2['test']
# 方法3
ws3 = wb2.get_sheet_by_name('test')
# 方法4
ws =wb2.active # 第一个工作表？


# 操作工作表
# 对sheet页设置一个颜色（16位的RGB颜色）
ws.sheet_properties.tabColor = 'ff72BA'



## 关于 行、列







### 关于单元格

#选择单个单元格
ws['A1']
ws.cell(1,1)

#选择单元格区域, 指针游走方向，从左到右，从上到下
ws['A1:B2']
ws['A1':'B2']

print(ws.values, end='\n\n')  # 生成器对象，将一行单元格作为元组单元--》组成的生成器
print(list(ws.values), end='\n\n') # 将生成器对象转换为列表数据，列表中是生成器中的所有数据


#
print(ws2['A1'].value)

cell = ws['A1']  # 获取指定位置的单元格对象
# 单元格列索引
print(cell.col_idx)
print(cell.column)
# 单元格行索引
print(cell.row)
# 单元格列名
print(cell.column_letter)
# 单元格的坐标
print(cell.coordinate)
# 单元格数字类型
# 默认是
# n：数值
# s：字符串
# d：日期时间
print(cell.data_type)
# 单元格编码格式，默认 utf-8
print(cell.encoding)
# 是否有样式
print(cell.has_style)  # 默认样式是 Normal，如果是默认样式，返回False
# 单元格样式
print(cell.style)
# 单元格样式id
print(cell.style_id)


#<generator object Worksheet._cells_by_row>
ws.iter_rows('A1:C2')
#<generator object Worksheet._cells_by_col>
ws.iter_cols('A1:C2')
