from openpyxl import load_workbook

wb = load_workbook("industry_name_code_number.xlsx")
sheet1 = wb["Sheet1"]

# code_num = sheet1["D"]
# for cell in code_num:
# print(cell.value)  # 对于是“公式”的单元格，cell.value是公式字符串

# 抽取小表行业代码
code_char_from = sheet1["B"]
code_char_to = [] # 目标表行业代码
for cell in code_char_from:
    code_char_to.append(cell.value)

# print(code_char_to) # 成功读取并形成
# print('053' in code_char_to)



wb = load_workbook("data_all.xlsx")
sheet1 = wb["all"]
#抽取出来大表中的行业代码
ind_code_from = sheet1["R"]
ind_code_to = []  # 大表中行业代码
for cell in ind_code_from:
    ind_code_to.append(cell.value)



# 判断 大表的每一行 是否 在备选的行业中，形成一个 长度为8025的数组
# 0-1数组
code_in = []

# 单独生成一个“备选行业代码”数组，然后填入到 data_all.xlsx中
# 如果“大表行业代码“在“小表行业代码列表”中，大表的“备选行业代码“就等于“大表行业代码”，否则
# 备选用代码
code_alt = []
#       每一个大表行业代码
for code in ind_code_to:
    if code in code_char_to:
        in_or_not = 1
        code_alt_use = code
    elif code[0:3] in code_char_to:
        in_or_not = 1
        code_alt_use = code[0:3]
    else:
        in_or_not = 0
        code_alt_use = 999999
    code_in.append(in_or_not)
    code_alt.append(code_alt_use)
# print(code_in)
# print(len(code_in)) # 8025说明数据正确
# print(code_alt)
# print(len(code_alt))


#把生成的0-1数组插入大表AG列
# 大表中尚空的AG列
iter_code_in = iter(code_in)
for cell in sheet1["AG"]:
    cell.value = next(iter_code_in)

# 把code_alt插入到AH列
iter_code_alt = iter(code_alt)
for cell in sheet1["AH"]:
    cell.value = next(iter_code_alt)

wb.save("data_all.xlsx")

